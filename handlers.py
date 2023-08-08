from aiogram import F, Router, types
from aiogram.filters import Command
from aiogram.types import Message

from aiogram import flags
from aiogram.fsm.context import FSMContext
from aiogram.types.callback_query import CallbackQuery
from states import Gen

import kb
import text
import states

from openpyxl import load_workbook

router = Router()
db = load_workbook('./gen.xlsx')
sheet = db.get_sheet_by_name('Sheet1')


@router.message(Command("start"))
async def start_handler(msg: Message):
    await msg.answer(text.greet.format(name=msg.from_user.full_name), reply_markup=kb.menu)

@router.message(F.text == "Меню")
@router.message(F.text == "Выйти в меню")
async def menu(msg: Message):
    await msg.answer(text.menu, reply_markup=kb.menu)

@router.callback_query(F.data == "menu")
async def menu(clbck: CallbackQuery, state: FSMContext):
    await clbck.message.answer(text.menu, reply_markup=kb.menu)

@router.callback_query(F.data == "anemo")
async def menu_anemo(clbck: CallbackQuery, state: FSMContext):
    await clbck.message.answer(text.menu_anemo, reply_markup=kb.menu_anemo)

@router.callback_query(F.data == "pyro")
async def menu_pyro(clbck: CallbackQuery, state: FSMContext):
    await clbck.message.answer(text.menu_pyro, reply_markup=kb.menu_pyro)

@router.callback_query(F.data == "cryo")
async def menu_cryo(clbck: CallbackQuery, state: FSMContext):
    await clbck.message.answer(text.menu_cryo, reply_markup=kb.menu_cryo)

@router.callback_query(F.data == "geo")
async def menu_anemo(clbck: CallbackQuery, state: FSMContext):
    await clbck.message.answer(text.menu_geo, reply_markup=kb.menu_geo)

@router.callback_query(F.data == "electro")
async def menu_geo(clbck: CallbackQuery, state: FSMContext):
    await clbck.message.answer(text.menu_electro, reply_markup=kb.menu_electro)

@router.callback_query(F.data == "dendro")
async def menu_dendro(clbck: CallbackQuery, state: FSMContext):
    await clbck.message.answer(text.menu_dendro, reply_markup=kb.menu_dendro)

@router.callback_query(F.data == "hydro")
async def menu_hydro(clbck: CallbackQuery, state: FSMContext):
    await clbck.message.answer(text.menu_hydro, reply_markup=kb.menu_hydro)

@router.callback_query(F.data == "back")
async def back(clbck: CallbackQuery, state: FSMContext):
    await clbck.message.answer(text.menu, reply_markup=kb.menu)
#Anemo
@router.callback_query(F.data == "kazuha")
async def menu_kazuha(clbck: CallbackQuery, state: FSMContext):
    await clbck.message.answer(text.menu_kazuha, reply_markup=kb.menu_kazuha)

@router.callback_query(F.data == "kazuha_ms")
async def menu_kazuha(clbck: CallbackQuery, state: FSMContext):
    argument = 2
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)
@router.callback_query(F.data == "kazuha_krit")
async def menu_kazuha(clbck: CallbackQuery, state: FSMContext):
    argument = 3
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "faruzan")
async def menu_faruzan(clbck: CallbackQuery, state: FSMContext):
    await clbck.message.answer(text.menu_faruzan, reply_markup=kb.menu_faruzan)

@router.callback_query(F.data == "faruzan_sup")
async def menu_faruzan(clbck: CallbackQuery, state: FSMContext):
    argument = 4
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)
@router.callback_query(F.data == "faruzan_dd")
async def menu_faruzan(clbck: CallbackQuery, state: FSMContext):
    argument = 5
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "heizou")
async def menu_heizou(clbck: CallbackQuery, state: FSMContext):
    argument = 6
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "venti")
async def menu_venti(clbck: CallbackQuery, state: FSMContext):
    await clbck.message.answer(text.menu_venti, reply_markup=kb.menu_venti)

@router.callback_query(F.data == "venti_ms")
async def menu_venti(clbck: CallbackQuery, state: FSMContext):
    argument = 7
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)
@router.callback_query(F.data == "venti_krit")
async def menu_venti(clbck: CallbackQuery, state: FSMContext):
    argument = 8
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "wanderer")
async def menu_wanderer(clbck: CallbackQuery, state: FSMContext):
    await clbck.message.answer(text.menu_wanderer, reply_markup=kb.menu_wanderer)

@router.callback_query(F.data == "wanderer_driver")
async def menu_wanderer(clbck: CallbackQuery, state: FSMContext):
    argument = 9
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "wanderer_dps")
async def menu_wanderer(clbck: CallbackQuery, state: FSMContext):
    argument = 10
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "jean")
async def menu_jean(clbck: CallbackQuery, state: FSMContext):
    await clbck.message.answer(text.menu_jean, reply_markup=kb.menu_jean)

@router.callback_query(F.data == "jean_dd")
async def menu_jean(clbck: CallbackQuery, state: FSMContext):
    argument = 11
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "jean_sup")
async def menu_jean(clbck: CallbackQuery, state: FSMContext):
    argument = 12
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "sayu")
async def menu_sayu(clbck: CallbackQuery, state: FSMContext):
    argument = 13
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "xiao")
async def menu_xiao(clbck: CallbackQuery, state: FSMContext):
    argument = 14
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "sucrose")
async def menu_sucrose(clbck: CallbackQuery, state: FSMContext):
    argument = 15
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "traveler_anemo")
async def menu_traveler_anemo(clbck: CallbackQuery, state: FSMContext):
    argument = 16
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "bennett")
async def menu_bennett(clbck: CallbackQuery, state: FSMContext):
    argument = 17
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "xiangling")
async def menu_xiangling(clbck: CallbackQuery, state: FSMContext):
    argument = 18
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "hu_tao")
async def menu_hu_tao(clbck: CallbackQuery, state: FSMContext):
    argument = 19
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "amber")
async def menu_amber(clbck: CallbackQuery, state: FSMContext):
    argument = 28
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "dehya")
async def menu_dehya(clbck: CallbackQuery, state: FSMContext):
    argument = 26
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "diluc")
async def menu_diluc(clbck: CallbackQuery, state: FSMContext):
    argument = 21
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "klee")
async def menu_klee(clbck: CallbackQuery, state: FSMContext):
    argument = 24
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "xinyan")
async def menu_xinyan(clbck: CallbackQuery, state: FSMContext):
    argument = 20
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "thoma")
async def menu_thoma(clbck: CallbackQuery, state: FSMContext):
    await clbck.message.answer(text.menu_thoma, reply_markup=kb.menu_thoma)

@router.callback_query(F.data == "thoma_supdd")
async def menu_thoma_supdd(clbck: CallbackQuery, state: FSMContext):
    argument = 23
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "thoma_shit")
async def menu_thoma_shit(clbck: CallbackQuery, state: FSMContext):
    argument = 22
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "yoimiya")
async def menu_yoimiya(clbck: CallbackQuery, state: FSMContext):
    argument = 25
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "yanfei")
async def menu_yanfei(clbck: CallbackQuery, state: FSMContext):
    argument = 27
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "ayaka")
async def menu_yanfei(clbck: CallbackQuery, state: FSMContext):
    argument = 69
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "chongyun")
async def menu_yanfei(clbck: CallbackQuery, state: FSMContext):
    argument = 67
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "diona")
async def menu_yanfei(clbck: CallbackQuery, state: FSMContext):
    argument = 68
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "eula")
async def menu_yanfei(clbck: CallbackQuery, state: FSMContext):
    argument = 66
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "ganyu")
async def menu_yanfei(clbck: CallbackQuery, state: FSMContext):
    argument = 64
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "kaeya")
async def menu_yanfei(clbck: CallbackQuery, state: FSMContext):
    argument = 73
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "layla")
async def menu_layla(clbck: CallbackQuery, state: FSMContext):
    await clbck.message.answer(text.menu_layla, reply_markup=kb.menu_layla)

@router.callback_query(F.data == "layla_supdd")
async def menu_layla_supdd(clbck: CallbackQuery, state: FSMContext):
    argument = 77
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "layla_support")
async def menu_layla_support(clbck: CallbackQuery, state: FSMContext):
    argument = 76
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "mika")
async def menu_mika(clbck: CallbackQuery, state: FSMContext):
    await clbck.message.answer(text.menu_mika, reply_markup=kb.menu_mika)

@router.callback_query(F.data == "mika_dd")
async def menu_mika_dd(clbck: CallbackQuery, state: FSMContext):
    argument = 72
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "mika_support")
async def menu_mika_support(clbck: CallbackQuery, state: FSMContext):
    argument = 71
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "qiqi")
async def menu_qiqi(clbck: CallbackQuery, state: FSMContext):
    argument = 70
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "rosara")
async def menu_rosara(clbck: CallbackQuery, state: FSMContext):
    argument = 65
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "shenhe")
async def menu_shenhe(clbck: CallbackQuery, state: FSMContext):
    argument = 74
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "aloy")
async def menu_aloy(clbck: CallbackQuery, state: FSMContext):
    argument = 88
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "albedo")
async def menu_albedo(clbck: CallbackQuery, state: FSMContext):
    argument = 79
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "gorou")
async def menu_gorou(clbck: CallbackQuery, state: FSMContext):
    argument = 83
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "itto")
async def menu_itto(clbck: CallbackQuery, state: FSMContext):
    argument = 78
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "noelle")
async def menu_noelle(clbck: CallbackQuery, state: FSMContext):
    argument = 80
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "traveler_geo")
async def menu_traveler_geo(clbck: CallbackQuery, state: FSMContext):
    argument = 87
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "ningguang")
async def menu_ningguang(clbck: CallbackQuery, state: FSMContext):
    await clbck.message.answer(text.menu_ningguang, reply_markup=kb.menu_ningguang)

@router.callback_query(F.data == "ningguang_dd")
async def menu_ningguang_dd(clbck: CallbackQuery, state: FSMContext):
    argument = 81
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "ningguang_supdd")
async def menu_ningguang_supdd(clbck: CallbackQuery, state: FSMContext):
    argument = 82
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "zhongli")
async def menu_zhongli(clbck: CallbackQuery, state: FSMContext):
    await clbck.message.answer(text.menu_zhongli, reply_markup=kb.menu_zhongli)

@router.callback_query(F.data == "zhongli_supdd")
async def menu_zhongli_supdd(clbck: CallbackQuery, state: FSMContext):
    argument = 86
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "zhongli_support")
async def menu_zhongli_support(clbck: CallbackQuery, state: FSMContext):
    argument = 85
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "yun_jin")
async def menu_yun_jin(clbck: CallbackQuery, state: FSMContext):
    argument = 84
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "beidou")
async def menu_beidou(clbck: CallbackQuery, state: FSMContext):
    argument = 53
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "cyno")
async def menu_cyno(clbck: CallbackQuery, state: FSMContext):
    argument = 60
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "dori")
async def menu_dori(clbck: CallbackQuery, state: FSMContext):
    argument = 59
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "fischl")
async def menu_fischl(clbck: CallbackQuery, state: FSMContext):
    argument = 51
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "keqing")
async def menu_keqing(clbck: CallbackQuery, state: FSMContext):
    argument = 52
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "traveler_electro")
async def menu_traveler_electro(clbck: CallbackQuery, state: FSMContext):
    argument = 63
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "kuki_shinobu")
async def menu_kuki_shinobu(clbck: CallbackQuery, state: FSMContext):
    argument = 57
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "lisa")
async def menu_lisa(clbck: CallbackQuery, state: FSMContext):
    argument = 62
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "raiden")
async def menu_raiden(clbck: CallbackQuery, state: FSMContext):
    argument = 56
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "razor")
async def menu_razor(clbck: CallbackQuery, state: FSMContext):
    argument = 58
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "sara")
async def menu_sara(clbck: CallbackQuery, state: FSMContext):
    await clbck.message.answer(text.menu_sara, reply_markup=kb.menu_sara)

@router.callback_query(F.data == "sara_supdd")
async def menu_sara_supdd(clbck: CallbackQuery, state: FSMContext):
    argument = 54
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "sara_support")
async def menu_sara_support(clbck: CallbackQuery, state: FSMContext):
    argument = 55
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "yae_miko")
async def menu_yae_miko(clbck: CallbackQuery, state: FSMContext):
    argument = 61
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "kirara")
async def menu_kirara(clbck: CallbackQuery, state: FSMContext):
    argument = 89
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "alhaitham")
async def menu_alhaitham(clbck: CallbackQuery, state: FSMContext):
    argument = 45
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "collei")
async def menu_collei(clbck: CallbackQuery, state: FSMContext):
    argument = 39
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "kaveh")
async def menu_kaveh(clbck: CallbackQuery, state: FSMContext):
    argument = 48
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "tighnari")
async def menu_tighnari(clbck: CallbackQuery, state: FSMContext):
    argument = 40
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "baizhu")
async def menu_baizhu(clbck: CallbackQuery, state: FSMContext):
    await clbck.message.answer(text.menu_baizhu, reply_markup=kb.menu_baizhu)

@router.callback_query(F.data == "baizhu_supdd")
async def menu_baizhu_supdd(clbck: CallbackQuery, state: FSMContext):
    argument = 50
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "baizhu_shit")
async def menu_baizhu_shit(clbck: CallbackQuery, state: FSMContext):
    argument = 49
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "traveler_dendro")
async def menu_traveler_dendro(clbck: CallbackQuery, state: FSMContext):
    await clbck.message.answer(text.menu_traveler_dendro, reply_markup=kb.menu_traveler_dendro)

@router.callback_query(F.data == "traveler_dendro_supdd")
async def menu_traveler_dendro_supdd(clbck: CallbackQuery, state: FSMContext):
    argument = 41
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "traveler_dendro_support")
async def menu_traveler_dendro_support(clbck: CallbackQuery, state: FSMContext):
    argument = 42
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "nahida")
async def menu_nahida(clbck: CallbackQuery, state: FSMContext):
    await clbck.message.answer(text.menu_nahida, reply_markup=kb.menu_nahida)

@router.callback_query(F.data == "nahida_support")
async def menu_nahida_support(clbck: CallbackQuery, state: FSMContext):
    argument = 46
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "nahida_dps")
async def menu_nahida_dps(clbck: CallbackQuery, state: FSMContext):
    argument = 47
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "yaoyao")
async def menu_yaoyao(clbck: CallbackQuery, state: FSMContext):
    await clbck.message.answer(text.menu_yaoyao, reply_markup=kb.menu_yaoyao)

@router.callback_query(F.data == "yaoyao_react")
async def menu_yaoyao_react(clbck: CallbackQuery, state: FSMContext):
    argument = 44
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "yaoyao_support")
async def menu_yaoyao_support(clbck: CallbackQuery, state: FSMContext):
    argument = 43
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "ayato")
async def menu_ayato(clbck: CallbackQuery, state: FSMContext):
    argument = 30
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "barbara")
async def menu_barbara(clbck: CallbackQuery, state: FSMContext):
    argument = 36
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "childe")
async def menu_childe(clbck: CallbackQuery, state: FSMContext):
    argument = 35
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "kokomi")
async def menu_kokomi(clbck: CallbackQuery, state: FSMContext):
    argument = 31
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "mona")
async def menu_mona(clbck: CallbackQuery, state: FSMContext):
    argument = 37
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "nilou")
async def menu_nilou(clbck: CallbackQuery, state: FSMContext):
    argument = 38
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "xingqiu")
async def menu_xingqiu(clbck: CallbackQuery, state: FSMContext):
    argument = 29
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "yelan")
async def menu_yelan(clbck: CallbackQuery, state: FSMContext):
    argument = 32
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "candace")
async def menu_candace(clbck: CallbackQuery, state: FSMContext):
    await clbck.message.answer(text.menu_candace, reply_markup=kb.menu_candace)

@router.callback_query(F.data == "candace_dd")
async def menu_candace_dd(clbck: CallbackQuery, state: FSMContext):
    argument = 33
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)

@router.callback_query(F.data == "candace_support")
async def menu_candace_support(clbck: CallbackQuery, state: FSMContext):
    argument = 34
    await clbck.message.answer(text.info.format(builld = sheet.cell(row=argument, column=2).value, set = sheet.cell(row=argument, column=3).value, flower = sheet.cell(row=argument, column=4).value, feather = sheet.cell(row=argument, column=5).value, watches = sheet.cell(row=argument, column=6).value, cup = sheet.cell(row=argument, column=7).value, crown = sheet.cell(row=argument, column=8).value, weapon= sheet.cell(row=argument, column=9).value, maybe= sheet.cell(row=argument, column=10).value), reply_markup=kb.back)
